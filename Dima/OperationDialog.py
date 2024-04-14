import mysql.connector
import datetime
import docx
import os
from docxtpl import DocxTemplate
from PyQt5 import QtCore, QtWidgets
from openpyxl.reader.excel import load_workbook

from Dima.AddDialog import Ui_AddDialog
from Dima.OperationDialogController import OperationDialogController
from support_file import SupportClass
from support_file import show_notification


# TODO: добавить обновление в MainWindow (попробовать с помощью commit'a)
#  сделать функцию принять товар чтобы пользователь вписывал
#  или выбирал дату поставки и количество товара добавить на товар
#  примерно тоже самое сделать для функции переместить товар
#  изменить передачу connection на cursor (везде)

class Ui_OperationDialog(object):
    def __init__(self, operation, connection, worker_name):

        self.db_data = {}
        self.original_data = {}
        self.worker_name = worker_name
        self.connection = connection
        self.cursor = self.connection.cursor()
        self.operation_type = operation
        self.table_name = "Current_product"
        self.document_word = None
        self.paths_txt = "C:\\Users\\user\\PycharmProjects\\project_vlada\\paths.txt"
        self.index_path_word = 6
        self.index_path_excel = 7

        self.operationDialog = QtWidgets.QDialog()
        self.operationDialog.setObjectName("operationDialog")
        self.operationDialog.resize(1200, 800)

        self.verticalLayoutWidget = QtWidgets.QWidget(self.operationDialog)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(30, 20, 1141, 751))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")

        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")

        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        self.find_line = QtWidgets.QLineEdit(self.verticalLayoutWidget)
        self.find_line.setInputMask("")
        self.find_line.setText("")
        self.find_line.setObjectName("find_line")
        self.find_line.textChanged.connect(lambda text: self.controller.support_instance.search_table(text))
        self.horizontalLayout.addWidget(self.find_line)

        self.choose_comboBox = QtWidgets.QComboBox(self.verticalLayoutWidget)
        self.choose_comboBox.setObjectName("choose_button")
        self.horizontalLayout.addWidget(self.choose_comboBox)

        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding,
                                           QtWidgets.QSizePolicy.Minimum)

        self.horizontalLayout.addItem(spacerItem)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.horizontalLayout.addLayout(self.horizontalLayout_3)

        self.cancel_button = QtWidgets.QPushButton(self.verticalLayoutWidget)
        self.cancel_button.setObjectName("cancel_button")
        self.horizontalLayout.addWidget(self.cancel_button)

        self.verticalLayout.addLayout(self.horizontalLayout)

        self.table_widget = QtWidgets.QTableWidget(self.verticalLayoutWidget)
        self.table_widget.setObjectName("table_widget")
        self.table_widget.setColumnCount(0)
        self.table_widget.setRowCount(0)
        self.table_widget.setEditTriggers(QtWidgets.QAbstractItemView.AllEditTriggers)
        self.table_widget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.verticalLayout.addWidget(self.table_widget)

        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")

        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")

        self.check_excel = QtWidgets.QCheckBox(self.verticalLayoutWidget)
        self.check_excel.setObjectName("check_excel")
        self.verticalLayout_2.addWidget(self.check_excel)

        self.check_word = QtWidgets.QCheckBox(self.verticalLayoutWidget)
        self.check_word.setObjectName("check_word")
        self.verticalLayout_2.addWidget(self.check_word)

        self.horizontalLayout_4.addLayout(self.verticalLayout_2)

        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding,
                                            QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem1)

        self.save_button = QtWidgets.QPushButton(self.verticalLayoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.save_button.sizePolicy().hasHeightForWidth())
        self.save_button.setSizePolicy(sizePolicy)
        self.save_button.setMinimumSize(QtCore.QSize(100, 20))
        self.save_button.setObjectName("save_button")
        self.horizontalLayout_4.addWidget(self.save_button)

        self.verticalLayout.addLayout(self.horizontalLayout_4)

        self.info_line = QtWidgets.QLineEdit(self.verticalLayoutWidget)

        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.info_line.sizePolicy().hasHeightForWidth())

        self.info_line.setSizePolicy(sizePolicy)
        self.info_line.setMinimumSize(QtCore.QSize(0, 60))
        self.info_line.setInputMask("")
        self.info_line.setText("")
        self.info_line.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignTop)
        self.info_line.setObjectName("info_line")

        self.verticalLayout.addWidget(self.info_line)

        self.retranslateUi(self.operationDialog)
        QtCore.QMetaObject.connectSlotsByName(self.operationDialog)

        self.support_instance = SupportClass(self.table_name, self.connection, self.table_widget)
        self.controller = OperationDialogController(self, self.support_instance, self.table_widget, self.info_line,
                                                    operation, connection, self.worker_name)
        self.show()
        # отображаем таблицу с данными
        self.display_current_product_data()
        self.table_widget.cellChanged.connect(self.handle_cell_changed)

        # определяем функции к кнопкам
        self.connect_buttons_events(self.operation_type)

    def retranslateUi(self, operationDialog):
        _translate = QtCore.QCoreApplication.translate
        operationDialog.setWindowTitle(_translate("operationDialog", "operationDialog"))
        self.find_line.setPlaceholderText(_translate("operationDialog", "Найти..."))
        self.cancel_button.setText(_translate("operationDialog", "Отмена"))
        self.check_excel.setText(_translate("operationDialog", "Создать файл Excel"))
        self.check_word.setText(_translate("operationDialog", "Создать файл Word"))
        self.save_button.setText(_translate("operationDialog", "Сохранить"))
        self.info_line.setPlaceholderText(_translate("operationDialog", "Ваши действия..."))

    def show(self):
        self.operationDialog.show()

    def load_paths(self):
        paths = []
        if os.path.exists(self.paths_txt):
            with open(self.paths_txt, 'r') as file:
                paths = [line.strip() for line in file if line.strip()]
        return paths

    def fetch_client_or_warehouse(self, operation_type, client_or_warehouse_id):
        if operation_type == 'Продать':
            self.cursor.execute("SELECT name FROM Client WHERE id=%s", (client_or_warehouse_id,))
            name_client = self.cursor.fetchone()
            return name_client[0]
        elif operation_type == 'Переместить':
            self.cursor.execute("SELECT name FROM Warehouse WHERE id=%s", (client_or_warehouse_id,))
            name_warehouse = self.cursor.fetchone()
            return name_warehouse[0]

    def fetch_price(self, product_id):
        self.cursor.execute("SELECT price FROM Product_property WHERE id=%s", (product_id,))
        price = self.cursor.fetchone()
        return price[0]

    def create_word_document(self, operation_type, worker_name, client_or_warehouse_id, selected_items):
        current_datetime = datetime.datetime.now()
        context = {
            'date_time': current_datetime,
            'person': worker_name
        }
        for selected_row_data in selected_items:
            if operation_type == 'Списать':
                context['warehouse'] = selected_row_data[6]
                self.index_path_word = 2
            elif operation_type == 'Продать':
                context['warehouse'] = selected_row_data[6]
                context['client'] = self.fetch_client_or_warehouse(operation_type, client_or_warehouse_id)
                self.index_path_word = 0
            elif operation_type == 'Переместить':
                context['warehouse'] = self.fetch_client_or_warehouse(operation_type, client_or_warehouse_id)
                self.index_path_word = 4

        paths_from_file = self.load_paths()
        path = paths_from_file[self.index_path_word]
        self.document_word = DocxTemplate(path)
        self.document_word.render(context)

        # Создаем таблицу один раз перед циклом
        if operation_type == 'Переместить':
            table = self.document_word.add_table(rows=1, cols=5)
        else:
            table = self.document_word.add_table(rows=1, cols=4)
        table.autofit = True
        hdr_cells = table.rows[0].cells
        hdr_cells[0].text = 'Наименование'
        hdr_cells[1].text = 'Количество'
        hdr_cells[2].text = 'Цена'

        if operation_type == 'Принять':
            hdr_cells[3].text = 'Склад, на который принят товар'
        elif operation_type == 'Переместить':
            hdr_cells[3].text = 'Склад, с которого перемещен товар'
            hdr_cells[4].text = 'Дата поставки'
        elif operation_type in ['Продать', 'Списать']:
            hdr_cells[3].text = 'Дата поставки'

        # Добавляем строки в таблицу в цикле
        for selected_row_data in selected_items:
            row_cells = table.add_row().cells
            row_cells[0].text = selected_row_data[2]
            row_cells[1].text = selected_row_data[4]
            row_cells[2].text = str(self.fetch_price(selected_row_data[0]))

            if operation_type == 'Принять':
                row_cells[3].text = selected_row_data[6]
            elif operation_type == 'Переместить':
                row_cells[3].text = selected_row_data[6]
                row_cells[4].text = selected_row_data[7]
            elif operation_type in ['Продать', 'Списать']:
                row_cells[3].text = selected_row_data[7]

        table.style = 'Table Grid'
        document_path = f"{operation_type}_{current_datetime.strftime('%Y-%m-%d_%H-%M-%S')}.docx"
        self.document_word.save(document_path)

        os.startfile(document_path)

    def create_excel_document(self, operation_type, worker_name, client_or_warehouse_id, selected_items):
        current_datetime = datetime.datetime.now()
        column_index = 2
        row_index = 9

        if operation_type == 'Списать':
            self.index_path_excel = 3
            row_index = 10
        elif operation_type == 'Продать':
            self.index_path_excel = 1
            row_index = 11
        elif operation_type == 'Переместить':
            self.index_path_excel = 5
            row_index = 11

        paths_from_file = self.load_paths()
        path = paths_from_file[self.index_path_excel]
        workbook = load_workbook(path)
        sheet = workbook.active

        # Заполнение таблицы данными
        for selected_row_data in selected_items:
            sheet.cell(row=row_index, column=column_index).value = selected_row_data[2]
            sheet.cell(row=row_index, column=column_index + 1).value = selected_row_data[4]
            sheet.cell(row=row_index, column=column_index + 2).value = self.fetch_price(selected_row_data[0])

            if operation_type == 'Принять':
                sheet.cell(row=row_index, column=column_index + 3).value = selected_row_data[6]
            elif operation_type == 'Переместить':
                sheet.cell(row=row_index, column=column_index + 3).value = selected_row_data[6]
                sheet.cell(row=row_index, column=column_index + 4).value = selected_row_data[7]
            elif operation_type in ['Продать', 'Списать']:
                sheet.cell(row=row_index, column=column_index + 3).value = selected_row_data[7]

            row_index += 1

        sheet.cell(row=5, column=3).value = current_datetime
        sheet.cell(row=6, column=3).value = worker_name

        for selected_row_data in selected_items:
            if operation_type == 'Списать':
                sheet.cell(row=7, column=3).value = selected_row_data[6]
            elif operation_type == 'Продать':
                sheet.cell(row=7, column=3).value = selected_row_data[6]
                sheet.cell(row=8, column=3).value = self.fetch_client_or_warehouse(
                    operation_type, client_or_warehouse_id)
            elif operation_type == 'Переместить':
                sheet.cell(row=7, column=3).value = self.fetch_client_or_warehouse(
                    operation_type, client_or_warehouse_id)

        # Сохранение документа Excel
        document_path = f"{operation_type}_{current_datetime.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        workbook.save(document_path)

        # Открытие документа
        os.startfile(document_path)

    def display_current_product_data(self):
        try:
            # SQL-запрос для получения данных о товарах на складе
            query_data = f"SELECT cp.id, cp.current_product_id, cp.quantity, cp.delivery_id, cp.warehouse_id, cp.delivery_date, pp.current_product_name " \
                         f"FROM {self.table_name} AS cp " \
                         f"JOIN Product_property AS pp ON cp.current_product_id = pp.id"

            # Выполнение запроса
            self.cursor.execute(query_data)
            result_data = self.cursor.fetchall()

            # Список для хранения имен операций
            operator_names = []
            # Список для хранения имен складов
            warehouse_names = []

            # Получение имен операций и складов
            for row in result_data:
                # Получение имени операции
                operation_id = row[3]
                self.cursor.execute(f"SELECT type FROM Operation WHERE id = {operation_id}")
                operator_name = self.cursor.fetchone()[0]
                operator_names.append(operator_name)

                # Получение имени склада
                warehouse_id = row[4]
                self.cursor.execute(f"SELECT name FROM Warehouse WHERE id = {warehouse_id}")
                warehouse_name = self.cursor.fetchone()[0]
                warehouse_names.append(warehouse_name)

            # Добавление нового столбца в заголовок таблицы
            operation_column_title = "Количество на " + self.operation_type.lower()
            headers = ['ID', 'ID_товара', 'Наименование', 'Количество на складе', operation_column_title, 'Поставки',
                       'Склад',
                       'Дата поставки']
            num_cols = len(headers)
            self.table_widget.setColumnCount(num_cols)
            self.table_widget.setHorizontalHeaderLabels(headers)

            # Заполнение таблицы данными
            self.table_widget.setRowCount(len(result_data))
            for row_index, row_data in enumerate(result_data):
                # Получение данных строки
                product_id, current_product_id, quantity, delivery_id, warehouse_id, delivery_date, product_name = row_data

                # Создание элементов QTableWidgetItem для каждого столбца
                item_id = QtWidgets.QTableWidgetItem(str(product_id))
                item_current_product_id = QtWidgets.QTableWidgetItem(str(current_product_id))
                item_name = QtWidgets.QTableWidgetItem(str(product_name))
                item_quantity = QtWidgets.QTableWidgetItem(str(quantity))
                item_operator = QtWidgets.QTableWidgetItem(operator_names[row_index])
                item_warehouse = QtWidgets.QTableWidgetItem(warehouse_names[row_index])
                item_delivery_date = QtWidgets.QTableWidgetItem(str(delivery_date))

                # Установка элементов в соответствующие ячейки
                self.table_widget.setItem(row_index, 0, item_id)
                self.table_widget.setItem(row_index, 1, item_current_product_id)
                self.table_widget.setItem(row_index, 2, item_name)
                self.table_widget.setItem(row_index, 3, item_quantity)
                self.table_widget.setItem(row_index, 4,
                                          QtWidgets.QTableWidgetItem(""))  # Пустая ячейка для нового столбца
                self.table_widget.setItem(row_index, 5, item_operator)
                self.table_widget.setItem(row_index, 6, item_warehouse)
                self.table_widget.setItem(row_index, 7, item_delivery_date)

                # Установка флагов для ячеек редактирования
                for column_index in range(num_cols):
                    item = self.table_widget.item(row_index, column_index)
                    if item and column_index == 4:  # Проверка, что это столбец с количеством на складе
                        item.setFlags(item.flags() | QtCore.Qt.ItemIsEditable)
                    else:
                        if item:
                            item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)  # Убираем флаг редактирования

        except mysql.connector.Error as e:
            # Обработка ошибок при выполнении SQL-запросов
            print("Ошибка при выполнении запроса:", e)

    def save_operation_data(self):
        """
        Сохраняет данные операции в базу данных и обновляет информацию о товарах.
        """
        # Получаем тип операции и текущее время
        selected_destination = None
        import datetime
        current_time = datetime.datetime.now().strftime("%Y-%m-%d")

        # Получаем идентификатор работника
        worker_id = self.controller.get_worker_id(self.worker_name)

        # Получаем выбранный элемент в зависимости от типа операции
        if self.operation_type in ["Продать", "Переместить"]:
            selected_destination = self.choose_comboBox.currentText()
        elif self.operation_type in ["Принять", "Списать"]:
            selected_destination = self.table_widget.item(self.table_widget.currentRow(),
                                                          6).text()

        # Получаем идентификатор клиента или склада
        client_or_warehouse_id = self.controller.get_client_or_warehouse_id(selected_destination)

        # Получаем выбранные товары для операции
        selected_items = self.controller.get_selected_items()
        if not selected_items:
            show_notification("Выберите товары для операции")
            return

        self.controller.add_or_delete_product()

        self.controller.insert_operation_data(self.operation_type, client_or_warehouse_id, worker_id, current_time,
                                              additional_characteristics=None, selected_items=selected_items)

        # Обновляем информацию о товарах и складе
        self.controller.update_product_information(selected_items, client_or_warehouse_id)

        # Проверяем состояние чекбокса для создания Word документа
        if self.check_word.isChecked():
            self.create_word_document(self.operation_type, self.worker_name, client_or_warehouse_id, selected_items)

        if self.check_excel.isChecked():
            self.create_excel_document(self.operation_type, self.worker_name, client_or_warehouse_id, selected_items)

        # После сохранения операции вызываем метод для отображения данных
        self.display_current_product_data()

    def populateComboBox(self):
        cursor = self.cursor
        self.choose_comboBox.clear()  # Очищаем combobox

        if self.operation_type == "Продать":
            # Запрос данных о клиентах из базы данных и заполнение comboBox
            query = "SELECT name FROM Client"
            cursor.execute(query)
            clients = cursor.fetchall()
            for client in clients:
                self.choose_comboBox.addItem(client[0])
            # Добавляем опцию "Добавить" после основных элементов
            self.choose_comboBox.addItem("Добавить...")
            self.choose_comboBox.activated.connect(self.handleAddItem)
            self.choose_comboBox.activated[str].connect(self.controller.get_client_or_warehouse_id)
        elif self.operation_type == "Переместить":
            # Запрос названий складов из базы данных и заполнение comboBox
            query = "SELECT name FROM Warehouse"
            cursor.execute(query)
            warehouses = cursor.fetchall()
            for warehouse in warehouses:
                self.choose_comboBox.addItem(warehouse[0])

    def connect_buttons_events(self, operation_type):
        from Vlada.MainWindow import UiMainWindow
        # mainWindow_instance = UiMainWindow(self.worker_name, self.connection)
        # Подключение событий к обработчикам
        if operation_type in ["Списать", "Принять"]:
            self.choose_comboBox.hide()
            self.save_button.clicked.connect(self.controller.add_or_delete_product)
        self.save_button.clicked.connect(self.save_operation_data)
        # self.cancel_button.clicked.connect(mainWindow_instance.cancel_deletion_row)
        if operation_type in ["Продать", "Переместить"]:
            self.populateComboBox()

    def handle_cell_changed(self, row, column):
        """
        Проверяет, есть ли данные в столбце с количеством
        :param row:
        :param column:
        :return:
        """
        if column == 4:  # Проверяем, что изменения произошли в столбце с количеством на операцию
            item = self.table_widget.item(row, column)
            if item:
                new_quantity = item.text()
                # Проверяем, что введены только цифры
                if not new_quantity.isdigit():
                    # # Выводим сообщение об ошибке
                    # QtWidgets.QMessageBox.warning(self.operationDialog, "Ошибка", "Введены некорректные данные. "
                    #                                                               "Пожалуйста, введите число.")
                    # Возвращаем старое значение
                    original_quantity = self.original_data.get((row, column), "")
                    item.setText(original_quantity)
                else:
                    # Проверяем, не превышает ли новое количество на складе
                    max_quantity = int(self.table_widget.item(row, 3).text())  # Количество на складе
                    if int(new_quantity) > max_quantity:
                        # Выводим сообщение об ошибке
                        QtWidgets.QMessageBox.warning(self.operationDialog, "Ошибка", "Введено количество больше, "
                                                                                      "чем имеется на складе.")
                        # Возвращаем старое значение
                        original_quantity = self.original_data.get((row, column), "")
                        item.setText(original_quantity)
                    else:
                        # Сохраняем новое значение
                        self.db_data[(row, column)] = new_quantity

    def handleAddItem(self, index):
        # Проверяем, была ли выбрана опция "Добавить"
        if index == self.choose_comboBox.count() - 1:
            add_dialog = Ui_AddDialog(self.operation_type)
            add_dialog.exec()
            Ui_OperationDialog.add_dialog_instance = add_dialog
            self.populateComboBox()


if __name__ == "__main__":
    import sys
    from PyQt5 import QtWidgets
    from Dima.OperationDialog import Ui_OperationDialog

    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='admin',
        database='warehouse'
    )
    app = QtWidgets.QApplication(sys.argv)
    ui = Ui_OperationDialog("Списать", conn, "Иван Иванов")
    sys.exit(app.exec_())
